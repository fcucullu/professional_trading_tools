# By Francisco Cucullu

def CopiarSolapa(RutaOrigen,SolapaOrigen,RutaDestino,SolapaDestino):
    import openpyxl  #Esta library hay que instalarla con 'pip intall openpyxl'
    #Se tiene que cambiar la extension por que openpyxl no lee archivos .xls

    Origen = openpyxl.load_workbook(RutaOrigen) 
    Destino = openpyxl.load_workbook(RutaDestino) 
    HojaOrigen =  Origen[SolapaOrigen]
    HojaDestino = Destino[SolapaDestino]
    
    
    columnas = 10      
    filas = 19
    
    #Crea una lista vacia de ese tamanio
    lista=[]
    for i in range(1,filas+1):
        lista.append([])
    
    #Copia del archivo base    
    for f in range(1,filas+1):
        for c in range(1, columnas+1):
                e=HojaOrigen.cell(row=f,column=c)
                lista[f-1].append(e.value)
                
    #Pega en el archivo destino
    for f in range(1,filas+1):
        for c in range(1, columnas+1):
                j=HojaDestino.cell(row=f,column=c)
                j.value=lista[f-1][c-1]
    
    Destino.save(RutaDestino)
    del Origen, Destino, HojaOrigen, HojaDestino

###############################################################################

def CrearRuta(nuevaruta):
    import os
    if not os.path.exists(nuevaruta): os.makedirs(nuevaruta)

###############################################################################    
    
def MoverArchivos(RutaOrigen, RutaDestino):
    import shutil
    try:
        shutil.move(RutaOrigen,RutaDestino)
    except:
        pass
    
############################################################################### 

def EjecutarMacro(WorkbookPath, NombreMacro):
    import os
    import win32com.client
    if os.path.exists(WorkbookPath):
        excel = win32com.client.Dispatch('Excel.Application')
        #excel.visible=1
        excel.Workbooks.Open(Filename = WorkbookPath, ReadOnly=1)
        excel.Application.Run(NombreMacro)
        excel.Workbooks(1).Close(SaveChanges=0)
        excel.Application.Quit()
    del excel

###############################################################################

def CambiarNombre(NombreViejo, NombreNuevo):
    import os
    os.rename(NombreViejo, NombreNuevo)

###############################################################################

def CambiarXLSporXLSX(NombreViejo):
    import win32com.client as win32
    import os
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(NombreViejo)

    wb.SaveAs(NombreViejo+"x", FileFormat = 51)    #FileFormat = 51 is for .xlsx extension
    wb.Close()                                     #FileFormat = 56 is for .xls extension
    excel.Application.Quit()
    os.remove(NombreViejo)
    
###############################################################################
    
def DescargarCNV(AÑO,MES,DIA):
    import webbrowser
    webbrowser.open("http://iolnet.invertir.local/iolnet/mvc/recursos/ReporteRGCNV624Excel?idContraparte=151&fecha="+AÑO+"-"+MES+"-"+DIA)
    
###############################################################################
    
def DetectaDia():
    import time
    DIA = time.strftime('%d')
    MES = time.strftime('%m')
    AÑO = time.strftime('%Y') 
    return DIA, MES, AÑO

###############################################################################

def DetectaHorario():
    import datetime
    now = datetime.datetime.now() 
    horario = now.strftime("%H:%M")
    return horario

###############################################################################

def DetectaHorarioSegundos():
    import datetime
    now = datetime.datetime.now() 
    horario = now.strftime("%H:%M:%S")
    return horario

###############################################################################
    
def ReconocerUsuario():
    import getpass
    return getpass.getuser()

##############################################################################

def AbrirArchivo(Ruta):
    import os
    os.startfile(Ruta)
    
##############################################################################
    
def EjecutarMacroYAABIERTO(WorkbookPath, NombreMacro):
    import os
    import win32com.client
    if os.path.exists(WorkbookPath):
        excel = win32com.client.Dispatch('Excel.Application')
        #excel.visible=1
        #excel.Workbooks.Open(Filename = WorkbookPath, ReadOnly=1)
        excel.Application.Run(NombreMacro)
        #excel.Workbooks(1).Close(SaveChanges=0)
        #excel.Application.Quit()
    del excel
    
###############################################################################
    
def CambiarDirectorio(NuevoDir):
    import os
    os.chdir(os.path.join(os.path.abspath(os.path.curdir),NuevoDir))

##############################################################################
    
def Virtualshell():
    import os
    os.system('pipenv shell')
    
###############################################################################
    
def CorrerOtroScript(Script):
    import os
    codigo = 'pipenv run python '+Script
    os.system(codigo)    
    
###############################################################################

def DeltaDateAyer(deltaDays):
    from datetime import datetime, timedelta
    ayer_ = datetime.today() - timedelta(days=1)
    fechaOrigen_ = datetime.today() - timedelta(days=deltaDays+1)
    
    ayer = ayer_.strftime('%Y') + '-' + ayer_.strftime('%m') + '-' + ayer_.strftime('%d')
    fechaOrigen = fechaOrigen_.strftime('%Y') + '-' + fechaOrigen_.strftime('%m') + '-' + fechaOrigen_.strftime('%d')
    
    return ayer, fechaOrigen

###############################################################################

def DeltaDateHoy(deltaDays):
    from datetime import datetime, timedelta
    hoy_ = datetime.today()
    fechaOrigen_ = datetime.today() - timedelta(days=deltaDays)
    
    hoy = hoy_.strftime('%Y') + '-' + hoy_.strftime('%m') + '-' + hoy_.strftime('%d')
    fechaOrigen = fechaOrigen_.strftime('%Y') + '-' + fechaOrigen_.strftime('%m') + '-' + fechaOrigen_.strftime('%d')
    
    return hoy, fechaOrigen

##############################################################################

def weekday():
    from datetime import datetime
    weekday = datetime.today().weekday()
    if weekday == 0:
        weekday = 'lunes'
    elif weekday == 1:
        weekday = 'martes'
    elif weekday == 2:
        weekday == 'miercoles'
    elif weekday == 3:
        weekday = 'jueves'
    elif weekday == 4:
        weekday = 'viernes'
    elif weekday == 5:
        weekday = 'sabado'
    elif weekday == 6:
        weekday = 'domingo'
        
    return weekday
    